# -*- coding: utf-8 -*-
"""
Evals fijos Valle/Sax — regresion del agente de ingenieria Antigravity.

CI (GitHub):  pytest tests/evals -m ci -v
Local full:   pytest tests/evals -v
              python scripts/run_evals.py --full

Casos definidos en opengravity/runtime/evals/engineering_cases.json
"""

from __future__ import annotations

import glob
import json
import os

import pytest

from ci_validate_tasks import validate_memory_index, validate_task_schema
from memory_semantic_search import search_similar
from validate_deliverables import validate_from_state

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
TASKS_ROOT = os.path.join(REPO_ROOT, "opengravity", "runtime", "tasks")
EVALS_JSON = os.path.join(REPO_ROOT, "opengravity", "runtime", "evals", "engineering_cases.json")
MEMORY_INDEX = os.path.join(REPO_ROOT, "opengravity", "runtime", "memory", "index.json")
GEO_DIR = os.path.join(REPO_ROOT, "opengravity", "runtime", "geo")
SAX_FOLDER = os.path.join(
    REPO_ROOT, "Proyectos", "estudio coordinacion protecciones", "nudo sax"
)
VALLE_FOLDER = os.path.join(
    REPO_ROOT, "Proyectos", "Colectora Valle", "estudio protecciones julio"
)


def _load_evals() -> dict:
    with open(EVALS_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def _case(case_id: str) -> dict:
    for item in _load_evals()["cases"]:
        if item["id"] == case_id:
            return item
    raise KeyError(case_id)


def _state_path(task_id: str) -> str:
    folder = task_id
    return os.path.join(TASKS_ROOT, folder, "state.json")


def _failed(results) -> list:
    return [r for r in results if not r.ok]


@pytest.mark.ci
def test_evl01_sax_ledger_schema():
    """EVL-01: PROT-SAX state.json coherente y validation passed."""
    task_dir = "PROT-SAX-0F-OMEXOM"
    path = _state_path(task_dir)
    errors = validate_task_schema(task_dir, path)
    assert errors == [], errors
    assert _failed(validate_from_state(path, schema_only=True)) == []


@pytest.mark.ci
def test_evl02_valle_ledger_schema():
    """EVL-02: PROT-VALLE state.json coherente y validation passed."""
    task_dir = "PROT-VALLE-0B-REV03"
    path = _state_path(task_dir)
    errors = validate_task_schema(task_dir, path)
    assert errors == [], errors
    assert _failed(validate_from_state(path, schema_only=True)) == []


@pytest.mark.ci
def test_evl03_memory_lessons_mem001_007():
    """EVL-03: Decision memory index con MEM-001..007 y archivos presentes."""
    errors = validate_memory_index()
    assert errors == [], errors
    with open(MEMORY_INDEX, "r", encoding="utf-8") as f:
        index = json.load(f)
    ids = {lesson["id"] for lesson in index.get("lessons") or []}
    expected = {f"MEM-{i:03d}" for i in range(1, 8)}
    assert expected.issubset(ids), f"faltan lecciones: {expected - ids}"


@pytest.mark.ci
def test_evl04_rag_mem001_merged_cells():
    """EVL-04: busqueda semantica prioriza MEM-001 (celdas fusionadas RCC)."""
    case = _case("EVL-04")
    hits, _backend = search_similar(case["query"], profile="engineering", top_k=3)
    assert hits, "sin hits semanticos"
    assert hits[0].doc.doc_id == case["expect_top_doc_id"]


@pytest.mark.ci
def test_evl05_rag_mem006_git_cleanup():
    """EVL-05: busqueda semantica prioriza MEM-006 (entregables post git cleanup)."""
    case = _case("EVL-05")
    hits, _backend = search_similar(case["query"], profile="engineering", top_k=3)
    assert hits, "sin hits semanticos"
    assert hits[0].doc.doc_id == case["expect_top_doc_id"]


@pytest.mark.ci
def test_evl06_geo_manifests_valid():
    """EVL-06: manifiestos geo Valle/Sax con bbox WGS84 valido."""
    for name in ("colectora_valle.geo.json", "nudo_sax.geo.json"):
        path = os.path.join(GEO_DIR, name)
        assert os.path.isfile(path), path
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        bbox = data.get("site_bbox_wgs84") or {}
        assert bbox["north"] > bbox["south"]
        assert bbox["east"] > bbox["west"]


@pytest.mark.local
def test_evl07_sax_deliverables_on_disk(skip_local):
    """EVL-07: entregables Sax existen y pasan validador completo."""
    path = _state_path("PROT-SAX-0F-OMEXOM")
    bad = _failed(validate_from_state(path, schema_only=False))
    assert bad == [], [(r.name, r.detail, r.error_code) for r in bad]


@pytest.mark.local
def test_evl08_valle_deliverables_on_disk(skip_local):
    """EVL-08: entregables Valle existen y pasan validador completo."""
    path = _state_path("PROT-VALLE-0B-REV03")
    bad = _failed(validate_from_state(path, schema_only=False))
    assert bad == [], [(r.name, r.detail, r.error_code) for r in bad]


def _read_sax_com_rows(xlsx_path: str) -> list[tuple[str, str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str, str]] = []
    for r in range(1, ws.max_row + 1):
        com = ws.cell(r, 2).value
        if not com or not str(com).upper().startswith("COM-"):
            continue
        status = str(ws.cell(r, 3).value or "").strip()
        rev05 = str(ws.cell(r, 9).value or "").strip()
        rows.append((str(com), status, rev05))
    return rows


@pytest.mark.local
def test_evl09_sax_rcc_rev05_format(skip_local):
    """EVL-09: RCC Rev.05 Sax — 12 COM, 6 OK, 6 abiertos con comentario."""
    import openpyxl

    case = _case("EVL-09")
    matches = glob.glob(os.path.join(SAX_FOLDER, case["rcc_glob"]))
    assert matches, f"no RCC Rev05 en {SAX_FOLDER}"
    rows = _read_sax_com_rows(matches[0])
    assert len(rows) == case["expect_com_count"]

    ok_count = sum(1 for _, _, rev in rows if rev.upper() == "OK")
    assert ok_count == case["expect_ok_count"]

    open_rows = [(c, s, r) for c, s, r in rows if s.lower().startswith("abierto")]
    assert len(open_rows) >= case["expect_open_with_comment_min"]
    for com, _status, rev in open_rows:
        assert rev and rev.upper() != "OK", f"{com} abierto debe tener comentario, no OK vacio"


@pytest.mark.local
def test_evl10_valle_verdict_rec_in_report(skip_local):
    """EVL-10: informe Valle contiene veredicto REC y solicitud Rev. 0C."""
    case = _case("EVL-10")
    matches = glob.glob(os.path.join(VALLE_FOLDER, case["report_glob"]))
    assert matches, f"informe no encontrado en {VALLE_FOLDER}"
    text = open(matches[0], "r", encoding="utf-8").read().upper()
    for needle in case["expect_verdict_substrings"]:
        assert needle.upper() in text, f"falta '{needle}' en informe Valle"


def test_eval_catalog_matches_tests():
    """Meta: engineering_cases.json tiene exactamente 10 casos EVL-01..10."""
    data = _load_evals()
    ids = [c["id"] for c in data["cases"]]
    assert len(ids) == 10
    assert ids[0] == "EVL-01" and ids[-1] == "EVL-10"
